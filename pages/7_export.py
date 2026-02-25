import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from exports.pdf_export import generate_pdf

st.set_page_config(page_title="Export & Proposal", layout="wide")
st.title("📤 Export & Proposal Generator")
st.markdown("Generate customer-ready deliverables from your analysis.")

if not st.session_state.get('parsed_data'):
    st.warning("Please complete the analysis before exporting.")
    st.stop()

if not st.session_state.get('scenario_results'):
    st.warning("Please complete the Scenario Builder before exporting.")
    st.stop()

parsed = st.session_state.parsed_data
current_tco = st.session_state.current_tco
scenario_results = st.session_state.scenario_results
selected_platforms = st.session_state.selected_platforms
years = st.session_state.assumptions.get('years', 3)
health = parsed.get('health', {})
discovery = st.session_state.get('discovery', {})
renewal_data = st.session_state.get('renewal_data', {})

override = st.session_state.get('recommendation_override')
if override:
    final_recommendation = override
else:
    scores = {}
    for platform, r in scenario_results.items():
        fit_score = r['fit']['fit_score']
        roi_score = min(r['roi_pct'] / 2, 50)
        scores[platform] = (fit_score * 0.6) + (roi_score * 0.4)
    final_recommendation = max(scores, key=scores.get)

st.divider()

# ── Export Settings ───────────────────────────────────────────────
st.subheader("Export Settings")
col1, col2 = st.columns(2)
with col1:
    customer_name = st.text_input("Customer Name", value="Customer")
    preparer_name = st.text_input("Prepared By", value="John Valentine")
with col2:
    st.markdown("**Recommendation**")
    st.info(f"Platform: **{final_recommendation}**")
    override_choice = st.selectbox(
        "Override Recommendation",
        ["Auto (Recommended)"] + selected_platforms,
    )
    if override_choice != "Auto (Recommended)":
        final_recommendation = override_choice
        st.session_state.recommendation_override = override_choice

st.divider()

# ── PDF Proposal ──────────────────────────────────────────────────
st.subheader("📄 Customer Proposal PDF")
st.caption("Select sections to include — toggle based on your audience.")

col_s1, col_s2 = st.columns(2)
with col_s1:
    inc_cover = st.checkbox("Cover Page", value=True)
    inc_exec = st.checkbox("Executive Summary", value=True)
    inc_env = st.checkbox("Current Environment Analysis", value=True)
    inc_tco = st.checkbox("TCO Comparison", value=True)
with col_s2:
    inc_rec = st.checkbox("Platform Recommendation", value=True)
    inc_reinvest = st.checkbox("Reinvestment Strategy", value=True)
    inc_roadmap = st.checkbox("5-Phase Journey Roadmap", value=True)
    inc_next = st.checkbox("Next Steps & Call to Action", value=True)

sections = []
if inc_cover: sections.append('cover')
if inc_exec: sections.append('executive_summary')
if inc_env: sections.append('environment')
if inc_tco: sections.append('tco')
if inc_rec: sections.append('recommendation')
if inc_reinvest: sections.append('reinvestment')
if inc_roadmap: sections.append('roadmap')
if inc_next: sections.append('next_steps')

if st.button("🚀 Generate Proposal PDF", type="primary"):
    with st.spinner("Building your proposal..."):
        try:
            pdf_buffer = generate_pdf(
                customer_name=customer_name,
                preparer_name=preparer_name,
                parsed=parsed,
                current_tco=current_tco,
                scenario_results=scenario_results,
                selected_platforms=selected_platforms,
                final_recommendation=final_recommendation,
                years=years,
                discovery=discovery,
                sections=sections,
            )
            st.download_button(
                label="⬇️ Download Proposal PDF",
                data=pdf_buffer,
                file_name=f"{customer_name.replace(' ', '_')}_Private_Cloud_Proposal.pdf",
                mime="application/pdf"
            )
            st.success("✅ Proposal generated successfully!")
        except Exception as e:
            st.error(f"Error generating PDF: {str(e)}")
            st.exception(e)

st.divider()

# ── Excel Model ───────────────────────────────────────────────────
st.subheader("📊 Detailed Excel Model")

def generate_excel():
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E79")
    highlight_fill = PatternFill("solid", start_color="E2EFDA")
    currency_format = '$#,##0'

    def style_header(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Executive Summary sheet
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1['A1'] = f"Private Cloud ROI & TCO — {customer_name}"
    ws1['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws1['A3'] = "Recommended Platform"
    ws1['B3'] = final_recommendation
    ws1['B3'].font = Font(bold=True)
    ws1['A4'] = "Analysis Period"
    ws1['B4'] = f"{years} Years"
    ws1['A5'] = "Prepared By"
    ws1['B5'] = preparer_name

    ws1['A7'] = "Environment Summary"
    style_header(ws1['A7'])
    ws1.merge_cells('A7:B7')
    env_rows = [
        ("Total VMs", parsed.get('total_vms', 0)),
        ("Powered On", parsed.get('powered_on_vms', 0)),
        ("Powered Off", parsed.get('powered_off_vms', 0)),
        ("Total Hosts", parsed.get('total_hosts', 0)),
        ("Total vCPU", parsed.get('total_vcpu', 0)),
        ("vCPU:pCPU Ratio", f"{parsed.get('vcpu_pcpu_ratio', 0)}:1"),
        ("Health Score", f"{parsed.get('health', {}).get('overall_pct', 0)}%"),
    ]
    for i, (label, value) in enumerate(env_rows):
        ws1[f'A{8+i}'] = label
        ws1[f'B{8+i}'] = value
        ws1[f'A{8+i}'].font = Font(bold=True)

    ws1['D7'] = f"Current {years}-Year TCO"
    style_header(ws1['D7'])
    ws1.merge_cells('D7:F7')
    tco_rows = [
        ("Hardware", current_tco['hardware_refresh']),
        ("Facilities", current_tco['facilities']),
        ("Labor", current_tco['fte']),
        ("Licensing", current_tco['licensing']),
        ("Support", current_tco['support']),
        ("TOTAL", current_tco['total']),
    ]
    for i, (label, value) in enumerate(tco_rows):
        ws1[f'D{8+i}'] = label
        ws1[f'F{8+i}'] = value
        ws1[f'F{8+i}'].number_format = currency_format
        if label == "TOTAL":
            ws1[f'D{8+i}'].font = Font(bold=True)
            ws1[f'F{8+i}'].font = Font(bold=True)
            ws1[f'D{8+i}'].fill = highlight_fill
            ws1[f'F{8+i}'].fill = highlight_fill

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['D'].width = 25
    ws1.column_dimensions['F'].width = 18

    # Scenario Comparison sheet
    ws2 = wb.create_sheet("Scenario Comparison")
    headers = ['Metric'] + selected_platforms
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    metrics = [
        (f'{years}-Year TCO', 'total'),
        ('Annual Average', 'annual_average'),
        ('Licensing', 'licensing'),
        ('Hardware', 'hardware_refresh'),
        ('Labor', 'fte'),
        ('Implementation', 'implementation'),
        ('Savings vs Current', 'savings'),
        ('ROI %', 'roi_pct'),
        ('Payback (months)', 'payback_months'),
        ('Fit Score', None),
    ]
    for row_idx, (label, key) in enumerate(metrics):
        ws2.cell(row=2+row_idx, column=1, value=label).font = Font(bold=True)
        for col_idx, platform in enumerate(selected_platforms):
            r = scenario_results[platform]
            value = r['fit']['fit_score'] if key is None else r.get(key, '')
            cell = ws2.cell(row=2+row_idx, column=2+col_idx, value=value)
            if key not in ['roi_pct', 'payback_months', None]:
                cell.number_format = currency_format

    for i in range(len(selected_platforms)+1):
        ws2.column_dimensions[get_column_letter(i+1)].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

if st.button("Generate Excel Model", type="secondary"):
    with st.spinner("Building Excel model..."):
        excel_buffer = generate_excel()
    st.download_button(
        label="⬇️ Download Excel Model",
        data=excel_buffer,
        file_name=f"{customer_name.replace(' ', '_')}_Private_Cloud_ROI.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.divider()

# ── Text Summary ──────────────────────────────────────────────────
st.subheader("📋 Quick Text Summary")
rec = scenario_results[final_recommendation]
summary = f"""PRIVATE CLOUD ANALYSIS — {customer_name}
Prepared by: {preparer_name} | {years}-Year Analysis

RECOMMENDATION: {final_recommendation}
- {years}-Year TCO: ${rec['total']:,.0f}
- Savings vs Current: ${rec['savings']:,.0f}
- ROI: {rec['roi_pct']}% | Payback: {rec['payback_months']} months

CURRENT STATE {years}-YEAR TCO: ${current_tco['total']:,.0f}
ENVIRONMENT: {parsed.get('total_vms',0)} VMs | {parsed.get('total_hosts',0)} Hosts | Health: {health.get('overall','N/A')}

PLATFORM SCORES:"""
for p in selected_platforms:
    r = scenario_results[p]
    marker = " ◀ RECOMMENDED" if p == final_recommendation else ""
    summary += f"\n• {p}{marker}: {r['fit']['fit_score']}/100 fit | ${r['total']:,.0f} TCO | ${r['savings']:,.0f} savings"

st.text_area("Summary", value=summary, height=300)
