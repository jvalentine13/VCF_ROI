import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Export", layout="wide")
st.title("📤 Export")
st.markdown("Generate customer-ready deliverables from your analysis.")

if not st.session_state.get('parsed_data'):
    st.warning("Please complete the analysis before exporting.")
    st.stop()

if not st.session_state.get('scenario_results'):
    st.warning("Please complete the Scenario Builder before exporting.")
    st.stop()

parsed = st.session_state.parsed_data
current_tco = st.session_state.current_tco
scenario_results = st.session_state.scenario_results
selected_platforms = st.session_state.selected_platforms
years = st.session_state.assumptions.get('years', 3)
health = parsed.get('health', {})

override = st.session_state.get('recommendation_override')
if override:
    final_recommendation = override
else:
    scores = {}
    for platform, r in scenario_results.items():
        fit_score = r['fit']['fit_score']
        roi_score = min(r['roi_pct'] / 2, 50)
        scores[platform] = (fit_score * 0.6) + (roi_score * 0.4)
    final_recommendation = max(scores, key=scores.get)

# Customer name input
st.subheader("Export Settings")
col1, col2 = st.columns(2)
with col1:
    customer_name = st.text_input("Customer Name", value="Customer")
with col2:
    preparer_name = st.text_input("Prepared By", value="")

st.divider()

# Excel export
st.subheader("📊 Detailed Excel Model")
st.caption("Full TCO model with all scenarios, assumptions, and calculations.")

def generate_excel():
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", start_color="1F4E79")
    subheader_fill = PatternFill("solid", start_color="2E75B6")
    subheader_font = Font(bold=True, color="FFFFFF")
    highlight_fill = PatternFill("solid", start_color="E2EFDA")
    currency_format = '$#,##0'
    pct_format = '0.0%'

    def style_header(cell, sub=False):
        cell.font = subheader_font if sub else header_font
        cell.fill = subheader_fill if sub else header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def set_col_widths(ws, widths):
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    # ── Sheet 1: Executive Summary ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"

    ws1['A1'] = f"Private Cloud ROI & TCO Analysis — {customer_name}"
    ws1['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws1.merge_cells('A1:F1')

    ws1['A2'] = f"Prepared by: {preparer_name}"
    ws1['A2'].font = Font(italic=True, size=11)
    ws1.merge_cells('A2:F2')

    ws1['A3'] = f"Analysis Period: {years} Years"
    ws1['A3'].font = Font(italic=True, size=11)
    ws1.merge_cells('A3:F3')

    # Environment summary
    ws1['A5'] = "Current Environment"
    style_header(ws1['A5'])
    ws1.merge_cells('A5:B5')

    env_data = [
        ("Total VMs", parsed.get('total_vms', 0)),
        ("Powered On VMs", parsed.get('powered_on_vms', 0)),
        ("Powered Off VMs", parsed.get('powered_off_vms', 0)),
        ("Total Hosts", parsed.get('total_hosts', 0)),
        ("Total Clusters", parsed.get('total_clusters', 0)),
        ("Total vCPU", parsed.get('total_vcpu', 0)),
        ("Total vRAM (GB)", parsed.get('total_vram_gb', 0)),
        ("Storage (GB)", parsed.get('total_storage_gb', 0)),
        ("vCPU:pCPU Ratio", f"{parsed.get('vcpu_pcpu_ratio', 0)}:1"),
        ("VM Density", f"{parsed.get('vm_density', 0)}/host"),
        ("Environment Health", parsed.get('health', {}).get('overall', 'N/A')),
    ]
    for i, (label, value) in enumerate(env_data):
        ws1[f'A{6+i}'] = label
        ws1[f'B{6+i}'] = value
        ws1[f'A{6+i}'].font = Font(bold=True)

    # Current TCO
    ws1['D5'] = f"Current State {years}-Year TCO"
    style_header(ws1['D5'])
    ws1.merge_cells('D5:F5')

    tco_data = [
        ("Hardware Refresh", current_tco['hardware_refresh']),
        ("Facilities & Power", current_tco['facilities']),
        ("Labor (FTE)", current_tco['fte']),
        ("Licensing", current_tco['licensing']),
        ("Support & Maintenance", current_tco['support']),
        ("TOTAL", current_tco['total']),
    ]
    for i, (label, value) in enumerate(tco_data):
        ws1[f'D{6+i}'] = label
        ws1[f'F{6+i}'] = value
        ws1[f'F{6+i}'].number_format = currency_format
        if label == "TOTAL":
            ws1[f'D{6+i}'].font = Font(bold=True)
            ws1[f'F{6+i}'].font = Font(bold=True)
            ws1[f'D{6+i}'].fill = highlight_fill
            ws1[f'F{6+i}'].fill = highlight_fill

    # Recommendation
    row = 20
    ws1[f'A{row}'] = "Platform Recommendation"
    style_header(ws1[f'A{row}'])
    ws1.merge_cells(f'A{row}:F{row}')

    row += 1
    ws1[f'A{row}'] = "Recommended Platform"
    ws1[f'B{row}'] = final_recommendation
    ws1[f'B{row}'].font = Font(bold=True, color="375623")

    rec = scenario_results[final_recommendation]
    summary_rows = [
        (f"{years}-Year TCO", rec['total']),
        ("Savings vs Current", rec['savings']),
        ("ROI %", rec['roi_pct'] / 100),
        ("Payback Period (months)", rec['payback_months']),
    ]
    for i, (label, value) in enumerate(summary_rows):
        ws1[f'A{row+1+i}'] = label
        ws1[f'B{row+1+i}'] = value
        ws1[f'A{row+1+i}'].font = Font(bold=True)
        if '$' in label or 'TCO' in label or 'Savings' in label:
            ws1[f'B{row+1+i}'].number_format = currency_format
        elif '%' in label:
            ws1[f'B{row+1+i}'].number_format = pct_format

    # Flags
    row += 8
    ws1[f'A{row}'] = "Environment Flags"
    style_header(ws1[f'A{row}'])
    ws1.merge_cells(f'A{row}:F{row}')
    for i, flag in enumerate(health.get('flags', [])):
        ws1[f'A{row+1+i}'] = flag

    set_col_widths(ws1, {'A': 30, 'B': 25, 'C': 5, 'D': 30, 'E': 5, 'F': 20})

    # ── Sheet 2: Scenario Comparison ────────────────────────────
    ws2 = wb.create_sheet("Scenario Comparison")

    ws2['A1'] = "Platform Scenario Comparison"
    ws2['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws2.merge_cells(f'A1:{get_column_letter(len(selected_platforms)+1)}1')

    headers = ['Metric'] + selected_platforms
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        style_header(cell, sub=(col > 1))

    metrics = [
        (f'{years}-Year TCO', 'total', currency_format),
        ('Annual Average', 'annual_average', currency_format),
        ('Licensing', 'licensing', currency_format),
        ('Support', 'support', currency_format),
        ('Hardware', 'hardware_refresh', currency_format),
        ('Facilities', 'facilities', currency_format),
        ('Labor', 'fte', currency_format),
        ('Implementation', 'implementation', currency_format),
        ('Savings vs Current', 'savings', currency_format),
        ('ROI %', 'roi_pct', '0.0"%"'),
        ('Payback (months)', 'payback_months', '0.0'),
        ('Fit Score', None, '0'),
        ('Effective Hosts', 'effective_hosts', '0'),
    ]

    for row_idx, (label, key, fmt) in enumerate(metrics):
        ws2.cell(row=4+row_idx, column=1, value=label).font = Font(bold=True)
        for col_idx, platform in enumerate(selected_platforms):
            r = scenario_results[platform]
            if key == 'fit_score' or key is None:
                value = r['fit']['fit_score']
            else:
                value = r.get(key, '')
            cell = ws2.cell(row=4+row_idx, column=2+col_idx, value=value)
            cell.number_format = fmt
            if platform == final_recommendation:
                cell.fill = highlight_fill

    set_col_widths(ws2, {get_column_letter(i+1): 25 for i in range(len(selected_platforms)+1)})

    # ── Sheet 3: Assumptions ────────────────────────────────────
    ws3 = wb.create_sheet("Assumptions")
    ws3['A1'] = "Model Assumptions"
    ws3['A1'].font = Font(bold=True, size=14, color="1F4E79")

    assumptions_data = [
        ("Analysis Period (Years)", years),
        ("FTE Count", st.session_state.assumptions.get('fte_count', 3)),
        ("Avg Host Cost ($)", st.session_state.assumptions.get('hardware', {}).get('avg_host_cost', 35000)),
        ("Hardware Refresh Cycle (Years)", st.session_state.assumptions.get('hardware', {}).get('refresh_cycle_years', 4)),
        ("Power Cost per kWh ($)", st.session_state.assumptions.get('hardware', {}).get('power_cost_per_kwh', 0.10)),
        ("Datacenter Cost per Host/Year ($)", st.session_state.assumptions.get('hardware', {}).get('datacenter_cost_per_host', 2000)),
        ("Fully Loaded FTE Cost ($)", st.session_state.assumptions.get('fte', {}).get('avg_fully_loaded_cost', 150000)),
    ]

    ws3['A3'] = "Parameter"
    ws3['B3'] = "Value"
    style_header(ws3['A3'], sub=True)
    style_header(ws3['B3'], sub=True)

    for i, (label, value) in enumerate(assumptions_data):
        ws3[f'A{4+i}'] = label
        ws3[f'B{4+i}'] = value
        ws3[f'A{4+i}'].font = Font(bold=True)

    set_col_widths(ws3, {'A': 35, 'B': 20})

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

if st.button("Generate Excel Model", type="primary"):
    with st.spinner("Building Excel model..."):
        excel_buffer = generate_excel()
    st.download_button(
        label="⬇️ Download Excel Model",
        data=excel_buffer,
        file_name=f"{customer_name.replace(' ', '_')}_Private_Cloud_ROI.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.success("Excel model ready!")

st.divider()

# Text summary export
st.subheader("📋 Executive Summary Text")
st.caption("Copy and paste into an email or document.")

rec = scenario_results[final_recommendation]
summary_text = f"""
PRIVATE CLOUD ROI & TCO ANALYSIS
{customer_name}
Prepared by: {preparer_name}
Analysis Period: {years} Years

CURRENT ENVIRONMENT
- Total VMs: {parsed.get('total_vms', 0)}
- Total Hosts: {parsed.get('total_hosts', 0)}
- vCPU:pCPU Ratio: {parsed.get('vcpu_pcpu_ratio', 0)}:1
- Environment Health: {health.get('overall', 'N/A')}

CURRENT STATE {years}-YEAR TCO: ${current_tco['total']:,.0f}
- Annual Average: ${current_tco['annual_average']:,.0f}

RECOMMENDATION: {final_recommendation}
- {years}-Year TCO: ${rec['total']:,.0f}
- Savings vs Current: ${rec['savings']:,.0f}
- ROI: {rec['roi_pct']}%
- Payback Period: {rec['payback_months']} months

PLATFORM COMPARISON
"""

for platform in selected_platforms:
    r = scenario_results[platform]
    marker = " ◀ RECOMMENDED" if platform == final_recommendation else ""
    summary_text += f"• {platform}{marker}: ${r['total']:,.0f} ({years}yr TCO), ${r['savings']:,.0f} savings\n"

summary_text += f"""
KEY ENVIRONMENT FLAGS
"""
for flag in health.get('flags', []):
    summary_text += f"• {flag}\n"

st.text_area("Executive Summary", value=summary_text.strip(), height=400)

st.divider()
st.success("✅ Analysis complete! Use the sidebar to navigate back to any page to adjust assumptions.")